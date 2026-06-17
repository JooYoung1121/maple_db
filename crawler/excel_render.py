"""엑셀(.xlsx) → 표 뷰 JSON + 원본 스타일 보존 HTML 변환 (openpyxl 기반, FastAPI 비의존).

병합셀(rowspan/colspan), 배경색, 글자색, 볼드, 가로정렬을 보존한다.
정보공유 게시판(api/routes/info_board.py)과 시드 스크립트에서 공용으로 사용.
"""
from __future__ import annotations

import html
import io
from typing import Optional

MAX_ROWS = 2000
MAX_COLS = 60


def argb_to_css(color) -> Optional[str]:
    """openpyxl Color -> #rrggbb (rgb 타입만, 투명/테마/인덱스는 None)."""
    if color is None or getattr(color, "type", None) != "rgb":
        return None
    rgb = color.rgb
    if not isinstance(rgb, str):
        return None
    if len(rgb) == 8:  # AARRGGBB
        if rgb[:2] == "00":
            return None
        rgb = rgb[2:]
    if len(rgb) != 6:
        return None
    return "#" + rgb.lower()


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def parse_excel(data: bytes):
    """returns (excel_json: dict, excel_html: str)."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheets = []
    html_parts: list[str] = []
    multi = len(wb.worksheets) > 1

    for ws in wb.worksheets:
        max_row = min(ws.max_row or 0, MAX_ROWS)
        max_col = min(ws.max_column or 0, MAX_COLS)
        if max_row == 0 or max_col == 0:
            continue

        span_map: dict[tuple[int, int], tuple[int, int]] = {}
        hidden: set[tuple[int, int]] = set()
        for rng in ws.merged_cells.ranges:
            r0, c0, r1, c1 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
            span_map[(r0, c0)] = (r1 - r0 + 1, c1 - c0 + 1)
            for rr in range(r0, r1 + 1):
                for cc in range(c0, c1 + 1):
                    if (rr, cc) != (r0, c0):
                        hidden.add((rr, cc))

        rows = []
        thtml = ['<table class="xl">']
        for r in range(1, max_row + 1):
            row_cells = []
            thtml.append("<tr>")
            for c in range(1, max_col + 1):
                if (r, c) in hidden:
                    row_cells.append({"hidden": True})
                    continue
                cell = ws.cell(row=r, column=c)
                rowspan, colspan = span_map.get((r, c), (1, 1))
                text = cell_text(cell.value)

                bg = None
                if getattr(cell.fill, "patternType", None) == "solid":
                    bg = argb_to_css(cell.fill.fgColor)
                color = argb_to_css(getattr(cell.font, "color", None))
                bold = bool(getattr(cell.font, "bold", False))
                align = getattr(cell.alignment, "horizontal", None) or None

                cd: dict = {"v": text}
                if rowspan > 1:
                    cd["rowspan"] = rowspan
                if colspan > 1:
                    cd["colspan"] = colspan
                if bg:
                    cd["bg"] = bg
                if color:
                    cd["color"] = color
                if bold:
                    cd["bold"] = True
                if align:
                    cd["align"] = align
                row_cells.append(cd)

                styles = []
                if bg:
                    styles.append(f"background-color:{bg}")
                if color:
                    styles.append(f"color:{color}")
                if bold:
                    styles.append("font-weight:700")
                styles.append(f"text-align:{align or 'left'}")
                attrs = f' style="{";".join(styles)}"'
                if rowspan > 1:
                    attrs += f' rowspan="{rowspan}"'
                if colspan > 1:
                    attrs += f' colspan="{colspan}"'
                safe = html.escape(text).replace("\n", "<br>")
                thtml.append(f"<td{attrs}>{safe}</td>")
            thtml.append("</tr>")
            rows.append(row_cells)
        thtml.append("</table>")

        sheets.append({"name": ws.title, "ncols": max_col, "rows": rows})
        if multi:
            html_parts.append(f'<h4 class="xl-sheet">{html.escape(ws.title)}</h4>')
        html_parts.append("".join(thtml))

    if not sheets:
        raise ValueError("빈 엑셀이거나 읽을 수 있는 시트가 없습니다.")
    return {"sheets": sheets}, "\n".join(html_parts)
