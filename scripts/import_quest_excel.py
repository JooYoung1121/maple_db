#!/usr/bin/env python3
"""
엑셀(퀘매.xlsx) 기반 퀘스트 DB 리빌딩 스크립트

엑셀을 마스터 데이터로 삼아:
1. 모든 퀘스트 is_mapleland=0 리셋
2. 엑셀 퀘스트를 DB에 매칭 → 업데이트 or INSERT
3. 기존 mapledb.kr 664건 중 엑셀에 없는 것 is_mapleland=1 유지 (별도 로직)
4. 검증 통계 출력
"""

import json
import os
import re
import sqlite3
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl 필요. pip install openpyxl")
    sys.exit(1)

# ── 설정 ──
EXCEL_PATH = os.environ.get(
    "EXCEL_PATH", "/Users/user/Downloads/퀘매.xlsx"
)
DB_PATH = os.environ.get(
    "DB_PATH",
    os.path.join(os.path.dirname(__file__), "..", "data", "maple.db"),
)

SHEET_TO_AREA = {
    "빅토리아 아일랜드": "빅토리아 아일랜드",
    "엘나스 산맥": "엘나스/아쿠아로드",
    "루더스 호수": "루디브리엄",
    "니할 사막": "무릉/니할사막",
    "리프레": "리프레",
    "세계여행": "세계여행",
}

QUEST_SHEETS = list(SHEET_TO_AREA.keys())
NEW_ID_START = 90001


# ── 유틸 함수 ──
def clean_name(raw: str) -> str:
    """퀘스트명에서 ▶┗┃ 접두사 및 (Lv.XX) 제거, 공백 트림"""
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"^[▶┗┃\s]+", "", s)
    # (Lv.XX), (~Lv.XX), (Lv. XX) 제거
    s = re.sub(r"\s*\(~?Lv\.?\s*\d+\)\s*", " ", s)
    return s.strip()


def is_chain(raw: str) -> bool:
    """▶┗┃ 접두사 여부"""
    if not raw:
        return False
    return bool(re.match(r"^[▶┗┃]", str(raw).strip()))


def extract_level_from_name(raw: str) -> int | None:
    """퀘스트명에서 (Lv.XX) or (~Lv.XX) 추출"""
    if not raw:
        return None
    m = re.search(r"\(~?Lv\.?\s*(\d+)\)", str(raw))
    return int(m.group(1)) if m else None


def detect_quest_type(name_raw: str, all_cols: list) -> str:
    """일일/반복/일반 판별 — 모든 컬럼을 검사"""
    text = " ".join(str(v) for v in [name_raw] + all_cols if v)
    if "일일" in text:
        return "반복"
    if "반복" in text:
        return "반복"
    return "일반"


def parse_exp(val) -> int:
    """EXP 보상 파싱"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(",", "").strip()
    m = re.search(r"(\d+)", s)
    return int(m.group(1)) if m else 0


def parse_meso(val) -> int:
    """메소 보상 파싱 (숫자만 있으면 메소)"""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace(",", "").strip()
    # 숫자로만 이루어진 경우 메소
    m = re.match(r"^(\d+)$", s)
    return int(m.group(1)) if m else 0


def collect_conditions(row_vals: list) -> str:
    """Col D~K (idx 3~10) 를 JSON으로 수집"""
    conditions = []
    for v in row_vals[3:11]:
        if v is not None:
            s = str(v).strip()
            if s:
                conditions.append(s)
    return json.dumps(conditions, ensure_ascii=False) if conditions else ""


def collect_rewards_text(row_vals: list) -> str:
    """Col M~N (idx 12~13) + 이후 보상 관련 텍스트"""
    parts = []
    for v in row_vals[12:17]:  # Col M ~ Q (넉넉히)
        if v is not None:
            s = str(v).strip()
            if s:
                parts.append(s)
    return " | ".join(parts) if parts else ""


def collect_tip(row_vals: list, max_col: int) -> str:
    """비고/TIP 컬럼 수집 (Col 17+)"""
    parts = []
    for v in row_vals[17:max_col]:
        if v is not None:
            s = str(v).strip()
            if s:
                parts.append(s)
    return " | ".join(parts) if parts else ""


# ── 1단계: 엑셀 파싱 ──
def parse_excel(excel_path: str) -> list[dict]:
    """엑셀에서 퀘스트 데이터 추출"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    all_quests = []

    for sheet_name in QUEST_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: 시트 '{sheet_name}' 없음, 건너뜀")
            continue

        ws = wb[sheet_name]
        area = SHEET_TO_AREA[sheet_name]
        max_col = ws.max_column

        prev_level = 0
        prev_location = ""

        for r in range(3, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]

            # 퀘스트명이 없으면 스킵
            name_raw = row_vals[2]  # Col C (idx 2)
            if not name_raw or not str(name_raw).strip():
                continue

            name_raw_str = str(name_raw).strip()
            name = clean_name(name_raw_str)
            if not name:
                continue

            # 레벨
            level_val = row_vals[0]  # Col A
            if level_val is not None and isinstance(level_val, (int, float)) and level_val > 0:
                prev_level = int(level_val)
            level_req = prev_level

            # (Lv.XX) 오버라이드
            level_from_name = extract_level_from_name(name_raw_str)
            if level_from_name is not None:
                level_req = level_from_name

            # 시작 장소
            loc = row_vals[1]  # Col B
            if loc and str(loc).strip():
                prev_location = str(loc).strip()
            start_location = prev_location

            # 퀘스트 조건
            quest_conditions = collect_conditions(row_vals)

            # EXP 보상 (Col L = idx 11)
            exp_reward = parse_exp(row_vals[11])

            # 아이템 보상 텍스트
            rewards_text = collect_rewards_text(row_vals)

            # 메소 보상 (Col N = idx 13, 숫자만이면 메소)
            meso_reward = parse_meso(row_vals[13])

            # 퀘스트 타입
            quest_type = detect_quest_type(name_raw_str, row_vals[3:max_col])

            # 체인 여부
            chain = is_chain(name_raw_str)

            # TIP/비고
            tip = collect_tip(row_vals, max_col)

            all_quests.append({
                "name": name,
                "name_raw": name_raw_str,
                "level_req": level_req,
                "area": area,
                "start_location": start_location,
                "quest_conditions": quest_conditions,
                "exp_reward": exp_reward,
                "meso_reward": meso_reward,
                "rewards_text": rewards_text,
                "quest_type": quest_type,
                "is_chain": chain,
                "tip": tip,
                "is_mapleland": 1,
                "sheet": sheet_name,
            })

    return all_quests


# ── 2단계: DB 매칭 & 업데이트 ──
def rebuild_db(quests: list[dict], db_path: str):
    """DB 리빌딩"""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 2-0. 통계용: 현재 상태
    before_ml = cur.execute("SELECT COUNT(*) FROM quests WHERE is_mapleland=1").fetchone()[0]
    before_total = cur.execute("SELECT COUNT(*) FROM quests").fetchone()[0]
    print(f"\n[기존 DB] 총 {before_total}건, is_mapleland=1: {before_ml}건")

    # 2-1. 모든 퀘스트 is_mapleland=0 리셋
    cur.execute("UPDATE quests SET is_mapleland = 0")
    print(f"[리셋] 모든 퀘스트 is_mapleland=0 으로 리셋")

    # 2-2. 기존 퀘스트 이름 인덱스 구축
    all_db_quests = cur.execute(
        "SELECT id, name, area, level_req, description FROM quests"
    ).fetchall()

    # 정확 매칭 맵 (name → list of rows)
    exact_map: dict[str, list] = {}
    for row in all_db_quests:
        n = row["name"].strip() if row["name"] else ""
        exact_map.setdefault(n, []).append(dict(row))

    # 매칭 통계
    stats = {"exact_match": 0, "like_match": 0, "new_insert": 0, "total": len(quests)}
    matched_db_ids = set()
    next_new_id = NEW_ID_START

    # 기존 최대 ID 확인
    max_existing_id = cur.execute("SELECT MAX(id) FROM quests").fetchone()[0] or 0
    if next_new_id <= max_existing_id:
        next_new_id = max_existing_id + 1

    # 2-pass: 1) 정확 매칭 먼저, 2) LIKE 매칭 (정확 매칭이 우선권 가짐)
    # Pass 1: 정확 매칭
    for q in quests:
        name = q["name"]
        q["_matched_row"] = None

        if name in exact_map:
            candidates = exact_map[name]
            area_match = [c for c in candidates if c["area"] == q["area"] and c["id"] not in matched_db_ids]
            if area_match:
                q["_matched_row"] = area_match[0]
            else:
                unmatched = [c for c in candidates if c["id"] not in matched_db_ids]
                if unmatched:
                    q["_matched_row"] = unmatched[0]

            if q["_matched_row"]:
                matched_db_ids.add(q["_matched_row"]["id"])
                stats["exact_match"] += 1

    # Pass 2: LIKE 매칭 (정확 매칭 안 된 것만)
    for q in quests:
        if q["_matched_row"]:
            continue
        name = q["name"]
        core = re.sub(r"\(.*?\)", "", name).strip()
        if len(core) >= 3:
            excluded = list(matched_db_ids) if matched_db_ids else [0]
            placeholders = ",".join("?" for _ in excluded)
            like_results = cur.execute(
                f"SELECT id, name, area, level_req, description FROM quests WHERE name LIKE ? AND id NOT IN ({placeholders})",
                [f"%{core}%"] + excluded,
            ).fetchall()
            if like_results:
                area_match = [dict(r) for r in like_results if r["area"] == q["area"]]
                if area_match:
                    q["_matched_row"] = area_match[0]
                else:
                    q["_matched_row"] = dict(like_results[0])
                matched_db_ids.add(q["_matched_row"]["id"])
                stats["like_match"] += 1

    # Pass 3: DB 반영
    for q in quests:
        name = q["name"]
        matched_row = q.pop("_matched_row", None)

        if matched_row:
            # 매칭됨 → 업데이트
            matched_db_ids.add(matched_row["id"])
            q["db_id"] = matched_row["id"]

            cur.execute(
                """UPDATE quests SET
                    area = ?,
                    level_req = ?,
                    exp_reward = ?,
                    meso_reward = CASE WHEN ? > 0 THEN ? ELSE meso_reward END,
                    quest_type = ?,
                    npc_start = CASE WHEN ? != '' THEN ? ELSE npc_start END,
                    is_mapleland = 1
                WHERE id = ?""",
                (
                    q["area"],
                    q["level_req"],
                    q["exp_reward"],
                    q["meso_reward"], q["meso_reward"],
                    q["quest_type"],
                    q["start_location"], q["start_location"],
                    matched_row["id"],
                ),
            )
        else:
            # 새로 INSERT
            stats["new_insert"] += 1
            q["db_id"] = next_new_id

            cur.execute(
                """INSERT INTO quests (
                    id, name, level_req, area, npc_start, exp_reward, meso_reward,
                    quest_type, is_mapleland, description
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)""",
                (
                    next_new_id,
                    name,
                    q["level_req"],
                    q["area"],
                    q["start_location"],
                    q["exp_reward"],
                    q["meso_reward"],
                    q["quest_type"],
                    q.get("rewards_text", ""),
                ),
            )
            next_new_id += 1

    # 2-3. 세계여행 area가 없으면 추가
    # (세계여행은 기존 DB에 없을 수 있음)

    conn.commit()
    print(f"\n[매칭 결과]")
    print(f"  엑셀 총 퀘스트: {stats['total']}건")
    print(f"  정확 매칭: {stats['exact_match']}건")
    print(f"  유사 매칭: {stats['like_match']}건")
    print(f"  신규 삽입: {stats['new_insert']}건")

    # 2-4. is_mapleland=1 총 건수 확인
    ml_count = cur.execute("SELECT COUNT(*) FROM quests WHERE is_mapleland=1").fetchone()[0]
    print(f"\n[현재 is_mapleland=1] {ml_count}건 (엑셀 {stats['total']}건 반영)")

    conn.close()
    return stats


# ── 3단계: 검증 ──
def verify(quests: list[dict], db_path: str):
    """검증 통계 출력"""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    print("\n" + "=" * 60)
    print("검증 결과")
    print("=" * 60)

    # area별 퀘스트 수 비교
    print("\n[area별 퀘스트 수 (is_mapleland=1)]")
    excel_area_counts = {}
    for q in quests:
        excel_area_counts[q["area"]] = excel_area_counts.get(q["area"], 0) + 1

    db_area_counts = {}
    rows = conn.execute(
        "SELECT area, COUNT(*) as cnt FROM quests WHERE is_mapleland=1 GROUP BY area ORDER BY area"
    ).fetchall()
    for r in rows:
        db_area_counts[r["area"]] = r["cnt"]

    print(f"  {'area':<20} {'엑셀':>6} {'DB':>6} {'일치':>6}")
    print(f"  {'-'*20} {'-'*6} {'-'*6} {'-'*6}")
    all_areas = sorted(set(list(excel_area_counts.keys()) + list(db_area_counts.keys())))
    for area in all_areas:
        ec = excel_area_counts.get(area, 0)
        dc = db_area_counts.get(area, 0)
        match = "OK" if ec == dc else f"DIFF({dc-ec:+d})"
        print(f"  {area:<20} {ec:>6} {dc:>6} {match:>6}")

    # 총 건수
    total_ml = conn.execute("SELECT COUNT(*) FROM quests WHERE is_mapleland=1").fetchone()[0]
    total_all = conn.execute("SELECT COUNT(*) FROM quests").fetchone()[0]
    print(f"\n  is_mapleland=1 총: {total_ml}건")
    print(f"  DB 전체: {total_all}건")

    # level_req > 0
    lv_count = conn.execute(
        "SELECT COUNT(*) FROM quests WHERE is_mapleland=1 AND level_req > 0"
    ).fetchone()[0]
    print(f"  level_req > 0: {lv_count}건")

    # exp_reward > 0
    exp_count = conn.execute(
        "SELECT COUNT(*) FROM quests WHERE is_mapleland=1 AND exp_reward > 0"
    ).fetchone()[0]
    print(f"  exp_reward > 0: {exp_count}건")

    # 각 area에서 5건 샘플
    print("\n[area별 샘플 (5건씩)]")
    for area in sorted(excel_area_counts.keys()):
        print(f"\n  --- {area} ---")
        samples = conn.execute(
            "SELECT id, name, level_req, exp_reward, quest_type FROM quests WHERE is_mapleland=1 AND area=? ORDER BY level_req, id LIMIT 5",
            (area,),
        ).fetchall()
        for s in samples:
            print(f"    [{s['id']}] Lv.{s['level_req']} {s['name']} (EXP:{s['exp_reward']}, type:{s['quest_type']})")

    conn.close()


# ── 메인 ──
def main():
    print("=" * 60)
    print("엑셀 기반 퀘스트 DB 리빌딩")
    print(f"  엑셀: {EXCEL_PATH}")
    print(f"  DB:   {DB_PATH}")
    print("=" * 60)

    # 1단계: 엑셀 파싱
    print("\n[1단계] 엑셀 파싱 중...")
    quests = parse_excel(EXCEL_PATH)
    print(f"  총 {len(quests)}건 파싱 완료")

    # 시트별 통계
    sheet_counts = {}
    for q in quests:
        sheet_counts[q["sheet"]] = sheet_counts.get(q["sheet"], 0) + 1
    for sn, cnt in sheet_counts.items():
        print(f"    {sn}: {cnt}건")

    # 2단계: DB 리빌딩
    print("\n[2단계] DB 리빌딩 중...")
    stats = rebuild_db(quests, DB_PATH)

    # 3단계: 검증
    verify(quests, DB_PATH)

    print("\n완료!")


if __name__ == "__main__":
    main()
