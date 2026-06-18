#!/usr/bin/env python3
"""정보공유 게시판 시드: 버닝서버(옛날메이플) 1~70렙 퀘스트 육성 루트 가이드.

원작자 HWP 자료(버닝서버 1~70 육성)를 기준으로 정정·반영한 버전을 엑셀(.xlsx)로 재현해
info_posts 에 등록/갱신한다. parse_excel() 로 표 뷰 JSON + 원본 스타일 HTML 을 함께 생성.

  python3 scripts/seed_info_board_quest_guide.py            # 비었을 때만 시드(유저 글 보호)
  python3 scripts/seed_info_board_quest_guide.py --update    # 기존 시드 글을 최신본으로 갱신
"""
from __future__ import annotations

import io
import json
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from crawler.excel_render import parse_excel  # noqa: E402

DB_PATH = Path(__file__).resolve().parent.parent / "data" / "maple.db"
TITLE = "옛날메이플 1~70렙 퀘스트 육성 루트 가이드"
TITLE_LIKE = "옛날메이플%육성 루트 가이드"
NICKNAME = "운영자"

HEADER_FILL = PatternFill("solid", fgColor="E3EAF6")
NORMAL_FILL = PatternFill("solid", fgColor="FFFFFF")
PINK_FILL = PatternFill("solid", fgColor="FCE4EC")
TEAL_FILL = PatternFill("solid", fgColor="B7E1D8")
REWARD_COLOR = "1A56C4"
MERGE_UP = "@MERGE_UP"  # 보상 셀을 위 행과 세로 병합

THIN = Side(style="thin", color="C9CFD8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (lv, region, quest, reward, kind)  kind: normal | party | boss | limit | note
ROWS = [
    ("1LV~7LV", "메이플 아일랜드", "★올 퀘스트★", "릴렉스 체어 / 마리아 머리띠 / 빅스 무기", "normal"),
    ("8LV", "리스항구", "헤네시스 걸어서 or 필(택시)", "9LV 찍어두기", "normal"),
    ("9LV", "헤네시스", "<페티트> 퀘스트", "딱 9LV 퀘스트 진행 후 몹 잡아서 10LV 업", "normal"),

    ("10LV~15LV", "세계여행 태국",
     "★전직X 초보자★\n초보자 세계여행 (비용 300메소)\n초보자 택시비(헤네) 페리온<>헤네 120메소\n<차이> 플로팅마켓 가입퀘",
     "장소: 플로팅 마켓 > 외딴집 > 동구 밖\nnpc: 차이 > 키드 > 폰 > 룽튭 > 푸야이리 > 지트\n/ 보상 경험치 1,550, 3000메소, 인기도3\n전직 마을 주문서 사두기", "normal"),
    ("10LV~15LV", "헤네시스",
     "<브루스> 내 딸을 찾고 싶어1\n<이얀> 내 딸을 찾고 싶어2",
     "<브루스> >> <이얀>\n<이얀> >> <브루스>\n/ 보상 경험치 200, 500메소", "normal"),
    ("10LV~15LV", "", "★10렙 경험치 0 시작시 퀘스트 완료 후 경험치 99퍼★", "", "note"),
    ("10LV~15LV", "헤네시스", "<브루스> 버섯 몬스터를 연구하는 이유",
     "주황버섯의 갓 40개, 버섯의 포자 10개\n/ 보상 빨간 포션 25개", "normal"),
    ("10LV~15LV", "전직 마을", "직업 전직",
     "수련 퀘스트 추천 사냥터\n-엘리니아 북쪽 필드\n-엘리니아 남쪽숲 나무던전1 –슬라임 굴\n/ 보상 물약", "normal"),
    ("10LV~15LV", "엘리니아", "직업 수련 퀘스트 진행 1 ~ 4", MERGE_UP, "normal"),
    ("10LV~15LV", "엘리니아", "<요정 윙> 숙제 좀 도와줘",
     "슬라임방울 10개, 나뭇가지 30개, 물컹물컹한 액체 30개\n/ 보상 파란포션 25개", "normal"),
    ("10LV~15LV", "커닝시티", "<이카루스> 심심해요 1 ~ 2",
     "나뭇가지 40개, 물컹물컹한 액체 40개\n/ 보상 빨간 포션 25개, 파란 포션 15개\n★11렙 커닝시티 고무신 구입 이속+2★", "normal"),

    ("10LV~20LV (파티 퀘스트)", "헤네시스", "월묘 파티 퀘스트",
     "15LV부터 보상방 입장가능\n/ 보상 드랍 템 팔아서 물약값 (상옵 장비 확인필요)", "party"),

    ("15LV~20LV", "엘리니아", "<카산드라> 의문의 작은알", "요정마르 펫 퀘스트\n/ 보상 달팽이 펫 3일", "normal"),
    ("15LV~20LV", "커닝시티 / 뉴리프시티", "뉴 리프 시티 환영 퀴즈 1 ~ 12",
     "왕복 1만메소 사용\n정답: 뉴리프시티-시간여행자-전송 시스템-비거 벤의 지하-위험한 변두리-이상한 스위치-장소로 전이됨-열릴 수도, 닫힐 수도-도적 왕자-보안관-보물 사냥꾼-모험 준비가 되면\n/ 보상 경험치 3,996, 엘릭서 5개", "normal"),
    ("15LV~20LV", "헤네시스", "헤네시스 동쪽 풀숲 (미니던전 돼지 농장)", "파사 or 솔플\n★펫 먹기 구입★", "normal"),
    ("15LV~20LV", "리스항구", "<존>의 분홍색 꽃 바구니", "끈기의 숲\n/ 보상 나사 30개", "normal"),
    ("15LV~20LV", "커닝시티", "<넬라>의 의뢰", "가입 1,000메소", "normal"),
    ("15LV~20LV", "슬리피우드", "위험한 던전 <초록버섯>",
     "99 퀘스트 / 보상 하얀 포션 50개\n999 퀘스트 / 보상 빨간 포션 50개", "normal"),
    ("15LV~20LV", "커닝시티", "<넬라><마파>의 의로", "초록버섯 갓 50개, 버블링 방울 50개\n/ 보상 레몬 50개", "normal"),
    ("15LV~20LV", "커닝시티", "<넬라><슈미>의 의뢰", "나사5개, 가공된 나무 5개\n/ 보상 케이크 100개", "normal"),
    ("15LV~20LV", "커닝시티", "<넬라><안드레아>의 의뢰", "옥토퍼스의 다리 100개, 맑은물 1개\n/ 보상 20렙 직업장갑", "normal"),

    ("18LV~29LV 레벨 제한 퀘스트", "페리온", "<만지>와 비밀조직",
     "다크 엑스텀프 10마리\n/ 보상 경험치 1,400, 아리안트 이동권 5번", "limit"),

    ("20LV", "직업 마을 / 쇼와 마을", "20렙 장비 구매", "★전신 쇼와 목욕타월 추천★", "normal"),
    ("20LV", "커닝시티", "추락주의", "파사 or 솔플", "normal"),
    ("20LV", "커닝시티", "<슈미>가 잃어버린 동전", "보상 마을주문서 랜덤 30개", "normal"),
    ("20LV", "헤네시스", "<마야>와 이상한 약",
     "오토퍼스의 다리 20개, 죽은 자의 부적 40개, 물컹물컹한 액체 50개, 나뭇잎 50개\n/ 보상 5,000메소, 갈색 삿갓", "normal"),
    ("20LV", "페리온", "<윈스턴> 화석을 찾아서 퀘",
     "(초록버섯) 나뭇잎 화석 100개, (엑스텀프) 동물의 뼈화석 100개\n/ 보상 경험치 4,500", "normal"),
    ("20LV", "페리온", "<연계 퀘> 화석상자 배달 퀘",
     "엘리니아 <베티>에게 화석상자를 배송\n보상 30,000 메소, 달걀 50개", "normal"),
    ("20LV", "페리온", "<윈스턴> 나무심기", "(고스텀프) 묘목 54개 / 보상 빨간 포션 30개", "normal"),
    ("20LV", "슬리피우드", "위험한 던전 <뿔버섯>",
     "99 퀘스트 / 보상 하얀 포션 100개\n999 퀘스트 / 보상 주황 포션 50개", "normal"),

    ("20LV 필드보스", "헤네시스", "<피아> 오래된 달팽이", "<피아>대화 >> <제이>대화\n/ 보상 경험치 5", "boss"),
    ("20LV 필드보스", "헤네시스", "<밍밍부인> 이야기의 주인공", "<밍밍부인>대화 >> <요정 윙>대화\n/ 보상 경험치 5", "boss"),
    ("20LV 필드보스", "엘리니아", "<요정 윙> 달팽이 사냥",
     "<요정 윙> 파란 달팽이의 껍질 10개, 빨간 달팽이의 껍질 10개, 달팽이의 껍질 10개\n/ 보상 경험치 100", "boss"),
    ("20LV 필드보스", "리스항구 / 헤네시스", "<피아> 소원을 들어주는 무지개색 달팽이 껍질",
     "리스항구 마노 사냥 무지개색 껍질 <피아>\n/ 보상 경험치 7500, 30,000 메소, 인기도 3", "boss"),

    ("21LV~30LV (파티 퀘스트)", "커닝시티", "커닝시티 파티 퀘스트", "보상 물약, 주문서, 원석, 경험치, 물컹신", "party"),

    ("21LV", "슬리피우드", "개미굴1, 개미굴2 (미니던전 버섯이 자라는 굴)", "파사", "normal"),

    ("22LV", "슬리피우드", "위험한 던전 <좀비버섯>",
     "99 퀘스트 / 보상 마나 엘릭서 50개\n999 퀘스트 / 보상 파란 포션 50개", "normal"),

    ("23LV", "헤네시스", "<카밀라>의 유리 구슬",
     "히든스트리트 유타네 돼지농장 <유타> 유리 구슬 전달\n/ 보상 케이크 30개 or 돼지도감", "normal"),

    ("25LV~30LV", "리스항구", "<제인>과 와일드보어",
     "뿔버섯의 갓 120개, 와일드 보어의 송곳니 100개\n/ 보상 폴암 60퍼 or 완드 60퍼", "normal"),
    ("25LV~30LV", "커닝시티", "<넬라> 소문난 의뢰",
     "주니어 네키의 가죽 50개, 뿔버섯의 갓 100개\n/ 보상 하얀 포션 100개", "normal"),
    ("25LV~30LV", "커닝시티", "<넬라><크리스>의 의뢰", "동물의 가죽 20개, 뻣뻣한 깃털 50개\n/ 보상 강철 1개, 청동 1개", "normal"),
    ("25LV~30LV", "커닝시티", "<넬라><만슈타인>의 의뢰", "이블아이의 꼬리 200개\n/ 보상 30렙 직업 신발", "normal"),
    ("25LV~30LV", "커닝시티", "<정체불명 그녀>의 부탁", "와일드보어 사냥 퀘스트 템 깨진 거울조각 20개\n/ 보상 경험치 2,300", "normal"),
    ("25LV~30LV", "커닝시티", "<정체불명 그녀>의 정체", "죽은 자의 부적 100개\n/ 보상 경험치 2,300, 허름한 망토", "normal"),
    ("25LV~30LV", "슬리피우드", "위험한 던전 <이블아이>",
     "99 퀘스트 / 보상 마나 엘릭서 100개\n999 퀘스트 / 보상 마나 엘릭서 50개", "normal"),
    ("25LV~30LV", "일본 (세계여행)", "일본: 카에데 성 성문 안", "파사 or 솔플", "normal"),

    ("18LV~29LV 레벨 제한 퀘스트", "엘리니아", "<하인즈> 요정 숲의 전령1",
     "보상 오르비스 행 티켓 1개, 리프레 행 티켓 1개, 리프레 책자", "limit"),
    ("18LV~29LV 레벨 제한 퀘스트", "페리온", "<만지> 사막으로... (29렙 제한 퀘스트 필수)",
     "<만지>클릭 아리안트 이동\n(\"선택\" 오르비스 이동 봉달이퀘 / 사헬지대 1~3 사냥 1업 후 30렙 마가티아 망토 퀘)", "limit"),
    ("18LV~29LV 레벨 제한 퀘스트", "오르비스", "<만지> 특별한 지령 (29렙 제한 퀘스트 선택)",
     "<만지> >> <봉달이>\n/ 보상 봉달이의 작은 상자", "limit"),

    ("30LV", "마가티아", "<브로커 한> 알카드노의 망토",
     "<브로커 한>정보비 1만메소 >> 알카드노 협회\n정답: 리튬, 하인즈, 알케스터\n/ 보상 알카드노의 망토 이속5 점프2 (교불)", "normal"),
    ("30LV", "슬리피우드", "[가운퀘] <시워언해> 비밀의 책의 단서", "<시워언해>대화", "normal"),
    ("30LV", "엘리니아", "[가운퀘] <로니> 배고픈 로니",
     "루팡의 바나나 50개, 특제 장어구이 1개, 신선한 우유 1개\n/ 보상 비밀의 책", "normal"),
    ("30LV", "헤네시스", "[가운퀘] <리나> 특제 장어구이",
     "돼지의 머리 5개, 커즈아이의 꼬리 50개\n/ 보상 특제 장어구이 1개", "normal"),
    ("30LV", "엘리니아", "[가운퀘] <요정 로웬> 신선한 우유",
     "다이아몬드 1개 (가공)\n/ 보상 신선한 우유 1개", "normal"),
    ("30LV", "엘리니아 / 슬리피우드", "[가운퀘] <시워언해> 되찾은 비밀의 책",
     "보상 10,000 메소, 가운", "normal"),
    ("30LV", "슬리피우드", "위험한 던전 <좀비버섯>",
     "99 퀘스트 / 보상 새벽이슬 3개 or 황혼이슬 3개 or 순록의 우유 5개 or 무기 주문서 60퍼 랜덤\n"
     "999 퀘스트 / 보상 순록의 우유 3 or 망토·무기 주문서 60퍼 랜덤", "normal"),
    ("30LV", "전직 마을", "2차 전직", "검은구슬 30개 / 직업 선택", "normal"),

    ("30LV~50LV (파티 퀘스트)", "루디브리엄", "<슈피겔만> 몬스터 카니발", "난동군 영상 참고", "party"),

    ("30LV~35LV", "루디브리엄", "에오스 탑 100층", "파사 or 솔플", "normal"),

    ("35LV~50LV (파티 퀘스트)", "루디브리엄", "루디브리엄 파티퀘스트",
     "보상 물약, 주문서, 원석, 경험치, 35판 달성 금이간안경", "party"),

    ("35LV~45LV", "엘나스", "오르비스 탑 1층", "봉달이퀘 물고기 파밍 솔플", "normal"),
    ("36LV~46LV", "루디브리엄", "시간의 길1", "파사", "normal"),
    ("40LV~50LV", "오르비스", "산책로", "파사 or 솔플", "normal"),
    ("40LV~55LV", "오르비스", "구름공원 4", "파사", "normal"),
    ("46LV~50LV", "루디브리엄", "시간의 길 4", "파사", "normal"),
    ("50LV~60LV", "무릉 / 마가티아 / 오르비스", "무릉: 초급 수련장, 마가티아: 연구소 C-1, 오르비스: 구름공원 6", "파사", "normal"),

    ("50LV~55LV (파티 퀘스트)", "오르비스", "오르비스 파티퀘스트", "경험치 파티 (퀘스트 완 X)", "party"),

    ("55LV~70LV", "엘나스", "죽숲 2, 3, 4, 차가운 벌판2", "파사", "normal"),
]

KIND_FILL = {"normal": NORMAL_FILL, "party": PINK_FILL, "boss": PINK_FILL,
             "limit": PINK_FILL, "note": TEAL_FILL}

CONTENT = (
    "버닝서버(옛날메이플) 1~70레벨 퀘스트 위주 육성 루트 정리입니다.\n"
    "표(엑셀) 뷰와 원본 스타일 보기로 모두 확인할 수 있어요.\n"
    "※ 원작자 자료 기준으로 정정 반영했습니다. 추가 정정 사항은 댓글로 알려주세요."
)


def build_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "1-70렙 육성"
    headers = ["LV", "지역", "퀘스트", "퀘스트 조건 / 보상"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 54

    reward_merge_pairs = []  # (top_row, bottom_row) 보상 세로병합
    r = 2
    for lv, region, quest, reward, kind in ROWS:
        if kind == "note":
            lv_fill = PINK_FILL if ("파티" in lv or "보스" in lv or "제한" in lv) else NORMAL_FILL
            a = ws.cell(row=r, column=1, value=lv); a.fill = lv_fill
            a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=r, column=2, value="").fill = NORMAL_FILL
            cc = ws.cell(row=r, column=3, value=quest)
            cc.fill = TEAL_FILL
            cc.font = Font(bold=True, color="00695C")
            cc.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=r, column=4, value="").fill = TEAL_FILL
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
            for ci in (1, 2, 3, 4):
                ws.cell(row=r, column=ci).border = BORDER
            r += 1
            continue

        lv_fill = PINK_FILL if kind in ("party", "boss", "limit") else NORMAL_FILL
        a = ws.cell(row=r, column=1, value=lv); a.fill = lv_fill
        a.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        a.font = Font(bold=True, color="C2185B" if lv_fill is PINK_FILL else "333333")
        b = ws.cell(row=r, column=2, value=region); b.fill = NORMAL_FILL
        b.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cq = ws.cell(row=r, column=3, value=quest); cq.fill = KIND_FILL[kind]
        cq.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if reward == MERGE_UP:
            # 보상 칸을 위 행과 세로 병합 (전직/수련 등)
            ws.cell(row=r, column=4, value=None).fill = NORMAL_FILL
            reward_merge_pairs.append((r - 1, r))
        else:
            dr = ws.cell(row=r, column=4, value=reward); dr.fill = NORMAL_FILL
            dr.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            dr.font = Font(color=REWARD_COLOR)
        for ci in (1, 2, 3, 4):
            ws.cell(row=r, column=ci).border = BORDER
        r += 1

    _merge_runs(ws, col=1, start_row=2, end_row=r - 1)
    for top, bottom in reward_merge_pairs:
        ws.merge_cells(start_row=top, start_column=4, end_row=bottom, end_column=4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _merge_runs(ws, col: int, start_row: int, end_row: int):
    run_start = start_row
    prev = ws.cell(row=start_row, column=col).value
    for rr in range(start_row + 1, end_row + 2):
        cur = ws.cell(row=rr, column=col).value if rr <= end_row else object()
        if rr > end_row or cur != prev:
            if rr - 1 > run_start:
                ws.merge_cells(start_row=run_start, start_column=col, end_row=rr - 1, end_column=col)
            run_start = rr
            prev = cur


def main() -> int:
    update = "--update" in sys.argv
    conn = sqlite3.connect(DB_PATH)
    try:
        count = conn.execute("SELECT COUNT(*) FROM info_posts").fetchone()[0]
    except sqlite3.OperationalError:
        print("info_posts 테이블 없음 — init_db 먼저 필요")
        conn.close()
        return 0

    existing = conn.execute(
        "SELECT id FROM info_posts WHERE title LIKE ? ORDER BY id LIMIT 1", (TITLE_LIKE,)
    ).fetchone()
    if not update and count > 0:
        print(f"info_posts 이미 {count}건 — 시드 건너뜀")
        conn.close()
        return 0

    data = build_xlsx()
    excel_json, excel_html = parse_excel(data)
    ej = json.dumps(excel_json, ensure_ascii=False)
    print(f"파싱 OK: 행 {len(excel_json['sheets'][0]['rows'])}, HTML {len(excel_html)}자")

    if update and existing:
        conn.execute(
            "UPDATE info_posts SET title=?, content=?, excel_filename=?, excel_json=?, excel_html=? WHERE id=?",
            [TITLE, CONTENT, "quest_route_1-70.xlsx", ej, excel_html, existing[0]],
        )
        conn.commit()
        print(f"[갱신] info_posts id={existing[0]} → {TITLE}")
    else:
        cur = conn.execute(
            "INSERT INTO info_posts (nickname, title, content, excel_filename, excel_json, excel_html) VALUES (?,?,?,?,?,?)",
            [NICKNAME, TITLE, CONTENT, "quest_route_1-70.xlsx", ej, excel_html],
        )
        conn.commit()
        print(f"[등록] info_posts id={cur.lastrowid}")
    conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
