#!/usr/bin/env python3
"""
퀘스트 데이터 완전 초기화 후 엑셀 파일만으로 재구축하는 스크립트.
기존 quests 테이블을 DROP하고 엑셀 419건만 INSERT합니다.
"""

import json
import os
import re
import sqlite3
import sys

import openpyxl

# ── 경로 설정 ──
EXCEL_PATH = "/Users/user/Downloads/퀘매.xlsx"
DB_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "maple.db")
DB_PATH = os.path.abspath(DB_PATH)

# ── 시트별 설정 ──
SHEET_CONFIG = {
    "빅토리아 아일랜드": {"area": "빅토리아 아일랜드", "note_col": 18, "tip_col": 19},
    "엘나스 산맥":       {"area": "엘나스/아쿠아로드",  "note_col": 21, "tip_col": None},
    "루더스 호수":       {"area": "루디브리엄",         "note_col": 19, "tip_col": None},
    "니할 사막":         {"area": "무릉/니할사막",       "note_col": 18, "tip_col": None},
    "리프레":            {"area": "리프레",              "note_col": 19, "tip_col": None},
    "세계여행":          {"area": "세계여행",            "note_col": 19, "tip_col": None},
}

# ── 배경색 → difficulty 매핑 ──
COLOR_DIFFICULTY = {
    "FFFFFF00": "추천",        # 노랑 = 클리어 추천
    "FF00FF00": "필수",        # 초록 = 클리어 필수 / 돈 되는 퀘스트
    "FFFF0000": "비추천",      # 빨강 = 클리어 비추천
    "FF00FFFF": "일일",        # 시안 = 일일/반복 퀘스트
    "FF4A86E8": "월드이동",    # 파랑 = 월드 이동 퀘스트
    "FFFF9900": "히든",        # 주황 = 히든/레벨 제한 퀘스트
    "FFFF00FF": "체인",        # 보라 = 장기 체인 퀘스트
}

# ── 새 스키마 ──
CREATE_TABLE_SQL = """
DROP TABLE IF EXISTS quests;
CREATE TABLE quests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    level_req INTEGER DEFAULT 0,
    area TEXT NOT NULL,
    start_location TEXT,
    quest_conditions TEXT,
    exp_reward INTEGER DEFAULT 0,
    meso_reward INTEGER DEFAULT 0,
    item_reward TEXT,
    extra_reward TEXT,
    note TEXT,
    tip TEXT,
    difficulty TEXT,
    is_chain INTEGER DEFAULT 0,
    chain_parent TEXT,
    quest_type TEXT DEFAULT '일반',
    is_mapleland INTEGER DEFAULT 1
);
"""

INSERT_SQL = """
INSERT INTO quests (
    name, level_req, area, start_location, quest_conditions,
    exp_reward, meso_reward, item_reward, extra_reward, note, tip,
    difficulty, is_chain, chain_parent, quest_type, is_mapleland
) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
"""


def safe_int(val, default=0):
    """float/str → int, None → default"""
    if val is None:
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def get_cell_bg_color(cell):
    """셀의 배경색 RGB 문자열을 반환 (예: 'FFFFFF00')"""
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.rgb:
        rgb = str(fill.fgColor.rgb)
        if rgb and rgb != "00000000":
            return rgb
    return None


def classify_difficulty(color):
    """배경색으로 난이도 분류"""
    if not color:
        return None
    return COLOR_DIFFICULTY.get(color)


def classify_quest_type(quest_name, difficulty, note_text):
    """퀘스트 유형 분류"""
    combined = (quest_name or "") + " " + (note_text or "")
    if "(일일)" in combined or "(반복)" in combined or "반복 퀘스트" in combined or "일일 퀘스트" in combined:
        return "반복"
    if difficulty == "히든":
        return "히든"
    if difficulty == "월드이동":
        return "월드이동"
    return "일반"


def parse_quest_name(raw_name):
    """
    퀘스트명에서 접두사 제거, (Lv.XX) 추출.
    Returns: (clean_name, is_chain, level_override)
    """
    name = str(raw_name).strip()
    is_chain = False

    # ▶, ┗, ┃ 접두사 체크 및 제거
    if name and name[0] in ("▶", "┗", "┃"):
        is_chain = True
        name = name[1:].strip()

    # (Lv.XX) 추출
    level_override = None
    lv_match = re.search(r"\(Lv\.?(\d+)\)", name)
    if lv_match:
        level_override = int(lv_match.group(1))
        # (Lv.XX) 부분은 이름에서 제거
        name = re.sub(r"\s*\(Lv\.?\d+\)\s*", " ", name).strip()

    return name, is_chain, level_override


def collect_conditions(ws, row):
    """Col D(4) ~ K(11): None이 아닌 값들을 리스트로"""
    conditions = []
    for c in range(4, 12):
        val = ws.cell(row, c).value
        if val is not None:
            s = str(val).strip()
            if s:
                # 숫자인 경우 int로 변환 시도
                try:
                    s = str(int(float(s)))
                except (ValueError, TypeError):
                    pass
                conditions.append(s)
    return conditions


def collect_rewards(ws, row, note_col):
    """
    Col M(13) ~ (note_col - 1): 아이템/메소 보상 수집.
    Returns: (meso_reward, item_parts, extra_parts)
    """
    meso_reward = 0
    item_parts = []
    extra_parts = []

    reward_cols = list(range(13, note_col))
    # Col 13(M) = 아이템 보상 텍스트
    # Col 14(N) = 메소(숫자) 또는 아이템 텍스트
    # Col 15+ = 추가 보상

    for idx, c in enumerate(reward_cols):
        val = ws.cell(row, c).value
        if val is None:
            continue

        if isinstance(val, (int, float)):
            int_val = int(val)
            if c == 14:
                # Col N: 숫자 = 메소 보상
                meso_reward = int_val
            elif c == 13:
                # Col M이 숫자인 경우는 거의 없지만 안전하게 처리
                item_parts.append(str(int_val))
            else:
                # Col O 이후의 숫자: 추가 보상의 일부 (수량 등)
                extra_parts.append(str(int_val))
        else:
            s = str(val).strip()
            if not s:
                continue
            if c == 13:
                item_parts.append(s)
            elif c == 14:
                # 텍스트: "or", "and" 등 구분자이거나 아이템명
                if s.lower() in ("or", "and"):
                    item_parts.append(s)
                else:
                    item_parts.append(s)
            else:
                extra_parts.append(s)

    return meso_reward, item_parts, extra_parts


def parse_sheet(wb, sheet_name, config):
    """시트 하나를 파싱하여 퀘스트 레코드 리스트 반환"""
    ws = wb[sheet_name]
    area = config["area"]
    note_col = config["note_col"]
    tip_col = config["tip_col"]

    records = []
    prev_level = 0
    prev_location = ""
    last_non_chain_name = ""

    for row in range(3, ws.max_row + 1):
        raw_name = ws.cell(row, 3).value
        if not raw_name or not str(raw_name).strip():
            continue

        # 레벨 (빈 셀 = 위 행 계승)
        level_val = ws.cell(row, 1).value
        if level_val is not None:
            prev_level = safe_int(level_val)
        level = prev_level

        # 시작 장소 (빈 셀 = 위 행 계승)
        loc_val = ws.cell(row, 2).value
        if loc_val is not None and str(loc_val).strip():
            prev_location = str(loc_val).strip()
        start_location = prev_location

        # 퀘스트명 파싱
        clean_name, is_chain, level_override = parse_quest_name(raw_name)
        if level_override is not None:
            level = level_override

        # 배경색 → difficulty
        bg_color = get_cell_bg_color(ws.cell(row, 3))
        difficulty = classify_difficulty(bg_color)

        # 퀘스트 조건
        conditions = collect_conditions(ws, row)
        quest_conditions = json.dumps(conditions, ensure_ascii=False) if conditions else None

        # EXP
        exp_reward = safe_int(ws.cell(row, 12).value)

        # 아이템/메소 보상
        meso_reward, item_parts, extra_parts = collect_rewards(ws, row, note_col)

        item_reward = " | ".join(item_parts) if item_parts else None
        extra_reward = " | ".join(extra_parts) if extra_parts else None

        # 비고
        note_val = ws.cell(row, note_col).value
        note = str(note_val).strip() if note_val and str(note_val).strip() else None

        # TIP
        tip = None
        if tip_col:
            tip_val = ws.cell(row, tip_col).value
            tip = str(tip_val).strip() if tip_val and str(tip_val).strip() else None

        # 체인 퀘스트 부모
        chain_parent = None
        if is_chain:
            chain_parent = last_non_chain_name
        else:
            last_non_chain_name = clean_name

        # 퀘스트 유형
        quest_type = classify_quest_type(clean_name, difficulty, note)

        records.append((
            clean_name,
            level,
            area,
            start_location,
            quest_conditions,
            exp_reward,
            meso_reward,
            item_reward,
            extra_reward,
            note,
            tip,
            difficulty,
            1 if is_chain else 0,
            chain_parent,
            quest_type,
        ))

    return records


def main():
    print(f"DB 경로: {DB_PATH}")
    print(f"엑셀 경로: {EXCEL_PATH}")

    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: 엑셀 파일이 없습니다: {EXCEL_PATH}")
        sys.exit(1)

    if not os.path.exists(DB_PATH):
        print(f"ERROR: DB 파일이 없습니다: {DB_PATH}")
        sys.exit(1)

    # 엑셀 열기 (배경색 읽으려면 data_only=False)
    print("\n엑셀 파일 로딩...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

    # DB 연결
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # 1. quests 테이블 DROP & CREATE
    print("\n[1/4] quests 테이블 초기화...")
    conn.executescript(CREATE_TABLE_SQL)
    print("  → quests 테이블 재생성 완료")

    # 2. entity_names_en에서 quest 관련 삭제
    print("\n[2/4] entity_names_en에서 quest 데이터 삭제...")
    try:
        deleted = conn.execute(
            "DELETE FROM entity_names_en WHERE entity_type = 'quest'"
        ).rowcount
        conn.commit()
        print(f"  → {deleted}건 삭제 완료")
    except Exception as e:
        print(f"  → entity_names_en 삭제 실패 (테이블 없을 수 있음): {e}")

    # 3. 시트별 파싱
    print("\n[3/4] 엑셀 시트 파싱...")
    all_records = []
    area_counts = {}
    difficulty_counts = {}

    for sheet_name, config in SHEET_CONFIG.items():
        records = parse_sheet(wb, sheet_name, config)
        all_records.extend(records)
        area_counts[config["area"]] = len(records)
        for rec in records:
            diff = rec[11]  # difficulty index
            if diff:
                difficulty_counts[diff] = difficulty_counts.get(diff, 0) + 1
        print(f"  {sheet_name} → {config['area']}: {len(records)}건")

    print(f"\n  총 파싱: {len(all_records)}건")

    # 4. INSERT
    print("\n[4/4] DB INSERT...")
    conn.executemany(INSERT_SQL, all_records)
    conn.commit()

    # 검증
    total = conn.execute("SELECT COUNT(*) FROM quests").fetchone()[0]
    print(f"\n  → INSERT 완료: {total}건")

    # 통계 출력
    print("\n" + "=" * 50)
    print("통계")
    print("=" * 50)

    print("\n[지역별 건수]")
    rows = conn.execute(
        "SELECT area, COUNT(*) as cnt FROM quests GROUP BY area ORDER BY cnt DESC"
    ).fetchall()
    for r in rows:
        print(f"  {r['area']}: {r['cnt']}건")

    print("\n[난이도별 건수]")
    rows = conn.execute(
        "SELECT difficulty, COUNT(*) as cnt FROM quests GROUP BY difficulty ORDER BY cnt DESC"
    ).fetchall()
    for r in rows:
        label = r["difficulty"] if r["difficulty"] else "(미분류)"
        print(f"  {label}: {r['cnt']}건")

    print("\n[퀘스트 유형별 건수]")
    rows = conn.execute(
        "SELECT quest_type, COUNT(*) as cnt FROM quests GROUP BY quest_type ORDER BY cnt DESC"
    ).fetchall()
    for r in rows:
        print(f"  {r['quest_type']}: {r['cnt']}건")

    print(f"\n[체인 퀘스트] {conn.execute('SELECT COUNT(*) FROM quests WHERE is_chain = 1').fetchone()[0]}건")
    print(f"[보상 있는 퀘스트] {conn.execute('SELECT COUNT(*) FROM quests WHERE exp_reward > 0 OR meso_reward > 0 OR item_reward IS NOT NULL').fetchone()[0]}건")

    # 샘플 출력
    print("\n[샘플 5건]")
    samples = conn.execute("SELECT id, name, level_req, area, difficulty, quest_type FROM quests LIMIT 5").fetchall()
    for s in samples:
        print(f"  #{s['id']} {s['name']} (Lv.{s['level_req']}) [{s['area']}] 난이도={s['difficulty']} 유형={s['quest_type']}")

    conn.close()
    print("\n완료!")


if __name__ == "__main__":
    main()
