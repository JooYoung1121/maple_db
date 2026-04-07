"""Excel export routes"""
import os
import io
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from crawler.db import get_connection

router = APIRouter()

EXPORT_QUERIES = {
    "items": "SELECT * FROM items ORDER BY id",
    "mobs": "SELECT * FROM mobs ORDER BY level, id",
    "maps": "SELECT * FROM maps ORDER BY id",
    "npcs": "SELECT * FROM npcs ORDER BY id",
    "quests": "SELECT * FROM quests ORDER BY level_req, id",
    "blog": "SELECT * FROM blog_posts ORDER BY id",
}

VALID_TYPES = set(EXPORT_QUERIES.keys())


def _check_admin_pw(pw: str):
    """Verify admin password."""
    admin_pw = os.environ.get("GAME_ADMIN_PASSWORD", "1004")
    if pw != admin_pw:
        raise HTTPException(status_code=403, detail="비밀번호가 틀립니다.")


def _style_header(ws):
    """Apply header styling to the first row."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _write_sheet(ws, rows, title=None):
    """Write rows to a worksheet with headers."""
    if title:
        ws.title = title
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row[col] for col in headers])
        _style_header(ws)
        # Auto-width (approximation)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header)) + 4
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50), min_col=col_idx, max_col=col_idx):
                for cell in row:
                    val_len = len(str(cell.value or ""))
                    if val_len > max_len:
                        max_len = val_len
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 50)
    else:
        ws.append(["데이터 없음"])


def _to_streaming(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export")
def export_data(
    type: str = Query(default="items"),
    format: str = Query(default="xlsx"),
):
    if type not in VALID_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid type. Choose from: {', '.join(sorted(VALID_TYPES))}",
        )
    if format != "xlsx":
        raise HTTPException(status_code=400, detail="Only xlsx format is supported")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Database unavailable")

    try:
        rows = conn.execute(EXPORT_QUERIES[type]).fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    _write_sheet(ws, rows, title=type)

    return _to_streaming(wb, f"{type}.xlsx")


# ── 퀘스트 전용 엑셀 내보내기 ──

_QUEST_COLUMNS = (
    "id, name, level_req, area, difficulty, quest_type, "
    "start_location, exp_reward, meso_reward, item_reward, "
    "extra_reward, note, tip, is_chain, chain_parent, is_mapleland"
)


@router.get("/export/quests")
def export_quests(
    format: str = Query(default="xlsx"),
    pw: str = Query(default=""),
):
    _check_admin_pw(pw)
    if format != "xlsx":
        raise HTTPException(status_code=400, detail="Only xlsx format is supported")

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Database unavailable")

    try:
        # Sheet1: 전체 퀘스트
        all_quests = conn.execute(
            f"SELECT {_QUEST_COLUMNS} FROM quests ORDER BY level_req, id"
        ).fetchall()

        # KR names
        kr_rows = conn.execute(
            "SELECT entity_id, name_en FROM entity_names_en WHERE entity_type='quest' AND source='kms'"
        ).fetchall()
        kr_map = {r["entity_id"]: r["name_en"] for r in kr_rows}

        def _enrich(rows):
            result = []
            for r in rows:
                d = dict(r)
                d["name_kr"] = kr_map.get(d["id"], "")
                # Reorder: ID, name, name_kr, ...
                ordered = {
                    "ID": d["id"],
                    "이름": d["name"],
                    "레벨": d["level_req"],
                    "지역": d.get("area") or "",
                    "난이도": d.get("difficulty") or "",
                    "유형": d.get("quest_type") or "",
                    "시작장소": d.get("start_location") or "",
                    "EXP보상": d.get("exp_reward") or 0,
                    "메소보상": d.get("meso_reward") or 0,
                    "아이템보상": d.get("item_reward") or "",
                    "추가보상": d.get("extra_reward") or "",
                    "비고": d.get("note") or "",
                    "TIP": d.get("tip") or "",
                    "체인": d.get("is_chain") or 0,
                    "체인부모": d.get("chain_parent") or "",
                }
                result.append(ordered)
            return result

        all_enriched = _enrich(all_quests)

        # Sheet2: 메이플랜드 퀘스트
        ml_quests = conn.execute(
            f"SELECT {_QUEST_COLUMNS} FROM quests WHERE is_mapleland = 1 ORDER BY level_req, id"
        ).fetchall()
        ml_enriched = _enrich(ml_quests)

        # Sheet3: 통계
        area_stats = conn.execute(
            "SELECT COALESCE(area, '(없음)') as 지역, COUNT(*) as 퀘스트수 FROM quests GROUP BY area ORDER BY COUNT(*) DESC"
        ).fetchall()
        type_stats = conn.execute(
            "SELECT COALESCE(quest_type, '(없음)') as 유형, COUNT(*) as 퀘스트수 FROM quests GROUP BY quest_type ORDER BY COUNT(*) DESC"
        ).fetchall()
        level_stats = conn.execute("""
            SELECT
                CASE
                    WHEN level_req <= 10 THEN '1-10'
                    WHEN level_req <= 20 THEN '11-20'
                    WHEN level_req <= 30 THEN '21-30'
                    WHEN level_req <= 40 THEN '31-40'
                    WHEN level_req <= 50 THEN '41-50'
                    WHEN level_req <= 60 THEN '51-60'
                    WHEN level_req <= 70 THEN '61-70'
                    WHEN level_req <= 80 THEN '71-80'
                    WHEN level_req <= 90 THEN '81-90'
                    WHEN level_req <= 100 THEN '91-100'
                    ELSE '100+'
                END as 레벨구간,
                COUNT(*) as 퀘스트수
            FROM quests GROUP BY 레벨구간 ORDER BY MIN(level_req)
        """).fetchall()

    finally:
        conn.close()

    wb = openpyxl.Workbook()

    # Sheet1
    ws1 = wb.active
    _write_sheet(ws1, [dict(zip(all_enriched[0].keys(), r.values())) for r in all_enriched] if all_enriched else [], title="퀘스트_전체")

    # Sheet2
    ws2 = wb.create_sheet()
    _write_sheet(ws2, [dict(zip(ml_enriched[0].keys(), r.values())) for r in ml_enriched] if ml_enriched else [], title="메이플랜드_퀘스트")

    # Sheet3: 통계
    ws3 = wb.create_sheet(title="데이터_통계")
    ws3.append(["=== 지역별 퀘스트 수 ==="])
    ws3.append(["지역", "퀘스트수"])
    for r in area_stats:
        ws3.append([r[0], r[1]])
    ws3.append([])
    ws3.append(["=== 유형별 퀘스트 수 ==="])
    ws3.append(["유형", "퀘스트수"])
    for r in type_stats:
        ws3.append([r[0], r[1]])
    ws3.append([])
    ws3.append(["=== 레벨별 퀘스트 수 ==="])
    ws3.append(["레벨구간", "퀘스트수"])
    for r in level_stats:
        ws3.append([r[0], r[1]])

    return _to_streaming(wb, "quests_export.xlsx")


# ── 전체 데이터 엑셀 내보내기 ──

@router.get("/export/all-data")
def export_all_data(
    pw: str = Query(default=""),
):
    _check_admin_pw(pw)

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Database unavailable")

    try:
        items = conn.execute("SELECT * FROM items ORDER BY id").fetchall()
        mobs = conn.execute("SELECT * FROM mobs ORDER BY level, id").fetchall()
        maps = conn.execute("SELECT * FROM maps ORDER BY id").fetchall()
        npcs = conn.execute("SELECT * FROM npcs ORDER BY id").fetchall()
        quests = conn.execute("SELECT * FROM quests ORDER BY level_req, id").fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()

    ws1 = wb.active
    _write_sheet(ws1, items, title="아이템")

    ws2 = wb.create_sheet()
    _write_sheet(ws2, mobs, title="몬스터")

    ws3 = wb.create_sheet()
    _write_sheet(ws3, maps, title="맵")

    ws4 = wb.create_sheet()
    _write_sheet(ws4, npcs, title="NPC")

    ws5 = wb.create_sheet()
    _write_sheet(ws5, quests, title="퀘스트")

    return _to_streaming(wb, "all_data_export.xlsx")


# ── 대시보드 통계 API ──

@router.get("/export/dashboard-stats")
def dashboard_stats(pw: str = Query(default="")):
    """관리자 대시보드용 통합 통계"""
    _check_admin_pw(pw)

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Database unavailable")

    try:
        # 퀘스트 통계
        total_quests = conn.execute("SELECT COUNT(*) FROM quests").fetchone()[0]
        ml_quests = conn.execute("SELECT COUNT(*) FROM quests WHERE is_mapleland = 1").fetchone()[0]
        rewarded = conn.execute(
            "SELECT COUNT(*) FROM quests WHERE exp_reward > 0 OR meso_reward > 0 OR item_reward IS NOT NULL"
        ).fetchone()[0]

        # 지역별
        area_rows = conn.execute(
            "SELECT COALESCE(area, '(없음)') as area, COUNT(*) as cnt FROM quests GROUP BY area ORDER BY cnt DESC"
        ).fetchall()
        area_dist = [{"area": r[0], "count": r[1]} for r in area_rows]

        # 이름 언어 분포
        kr_count = conn.execute(
            "SELECT COUNT(*) FROM quests WHERE id IN (SELECT entity_id FROM entity_names_en WHERE entity_type='quest' AND source='kms')"
        ).fetchone()[0]
        # 한국어 이름 패턴: 한글 포함
        kr_name_count = conn.execute(
            "SELECT COUNT(*) FROM quests WHERE name GLOB '*[가-힣]*'"
        ).fetchone()[0]
        en_name_count = total_quests - kr_name_count
        name_lang = {"korean": kr_name_count, "english": en_name_count, "has_kr_translation": kr_count}

        # 레벨 분포
        level_rows = conn.execute("""
            SELECT
                CASE
                    WHEN level_req <= 10 THEN '1-10'
                    WHEN level_req <= 20 THEN '11-20'
                    WHEN level_req <= 30 THEN '21-30'
                    WHEN level_req <= 40 THEN '31-40'
                    WHEN level_req <= 50 THEN '41-50'
                    WHEN level_req <= 60 THEN '51-60'
                    WHEN level_req <= 70 THEN '61-70'
                    WHEN level_req <= 80 THEN '71-80'
                    WHEN level_req <= 90 THEN '81-90'
                    WHEN level_req <= 100 THEN '91-100'
                    ELSE '100+'
                END as range,
                COUNT(*) as cnt
            FROM quests GROUP BY range ORDER BY MIN(level_req)
        """).fetchall()
        level_dist = [{"range": r[0], "count": r[1]} for r in level_rows]

        # 데이터 품질
        area_null = conn.execute("SELECT COUNT(*) FROM quests WHERE area IS NULL OR area = ''").fetchone()[0]
        level_zero = conn.execute("SELECT COUNT(*) FROM quests WHERE level_req = 0 OR level_req IS NULL").fetchone()[0]
        has_reward = rewarded
        quality = {
            "area_null_count": area_null,
            "area_null_pct": round(area_null / total_quests * 100, 1) if total_quests > 0 else 0,
            "level_zero_count": level_zero,
            "level_zero_pct": round(level_zero / total_quests * 100, 1) if total_quests > 0 else 0,
            "has_reward_count": has_reward,
            "has_reward_pct": round(has_reward / total_quests * 100, 1) if total_quests > 0 else 0,
        }

        # 전체 데이터 현황
        entity_counts = {
            "items": conn.execute("SELECT COUNT(*) FROM items").fetchone()[0],
            "mobs": conn.execute("SELECT COUNT(*) FROM mobs").fetchone()[0],
            "maps": conn.execute("SELECT COUNT(*) FROM maps").fetchone()[0],
            "npcs": conn.execute("SELECT COUNT(*) FROM npcs").fetchone()[0],
            "quests": total_quests,
        }

        # 크롤링 상태
        crawl_status = {}
        for entity in ["items", "mobs", "maps", "npcs", "quests"]:
            try:
                crawled = conn.execute(
                    f"SELECT COUNT(*) FROM {entity} WHERE last_crawled_at IS NOT NULL"
                ).fetchone()[0]
                total = entity_counts[entity]
                latest = conn.execute(
                    f"SELECT MAX(last_crawled_at) FROM {entity}"
                ).fetchone()[0]
                crawl_status[entity] = {
                    "crawled": crawled,
                    "total": total,
                    "pct": round(crawled / total * 100, 1) if total > 0 else 0,
                    "latest": latest,
                }
            except Exception:
                crawl_status[entity] = {"crawled": 0, "total": 0, "pct": 0, "latest": None}

    finally:
        conn.close()

    return {
        "quest_stats": {
            "total": total_quests,
            "mapleland": ml_quests,
            "rewarded": rewarded,
        },
        "area_distribution": area_dist,
        "name_language": name_lang,
        "level_distribution": level_dist,
        "quality": quality,
        "entity_counts": entity_counts,
        "crawl_status": crawl_status,
    }


@router.get("/export/all-quests")
def export_all_quests_json(
    pw: str = Query(default=""),
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, ge=1, le=500),
    q: str = Query(default=""),
    area: str = Query(default=""),
    category: str = Query(default=""),
    is_mapleland: str = Query(default="all"),
):
    """관리자 대시보드용 전체 퀘스트 목록 (JSON)"""
    _check_admin_pw(pw)

    conditions = []
    params: list = []

    if q:
        conditions.append(
            "(name LIKE ? OR id IN (SELECT entity_id FROM entity_names_en WHERE entity_type='quest' AND name_en LIKE ?))"
        )
        params.extend([f"%{q}%", f"%{q}%"])
    if area:
        conditions.append("area = ?")
        params.append(area)
    if category:
        conditions.append("difficulty = ?")
        params.append(category)
    if is_mapleland == "1":
        conditions.append("is_mapleland = 1")
    elif is_mapleland == "0":
        conditions.append("(is_mapleland = 0 OR is_mapleland IS NULL)")

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    offset = (page - 1) * per_page

    try:
        conn = get_connection()
    except Exception:
        raise HTTPException(status_code=503, detail="Database unavailable")

    try:
        total = conn.execute(f"SELECT COUNT(*) FROM quests {where}", params).fetchone()[0]
        rows = conn.execute(
            f"SELECT id, name, level_req, area, difficulty, quest_type, start_location, "
            f"exp_reward, meso_reward, item_reward, is_mapleland "
            f"FROM quests {where} ORDER BY level_req, id LIMIT ? OFFSET ?",
            params + [per_page, offset],
        ).fetchall()

        results = [dict(r) for r in rows]
        quest_ids = [q["id"] for q in results]
        kr_map = {}
        if quest_ids:
            placeholders = ",".join("?" for _ in quest_ids)
            kr_rows = conn.execute(
                f"SELECT entity_id, name_en FROM entity_names_en WHERE entity_type='quest' AND source='kms' AND entity_id IN ({placeholders})",
                quest_ids,
            ).fetchall()
            kr_map = {r["entity_id"]: r["name_en"] for r in kr_rows}

        for quest in results:
            quest["name_kr"] = kr_map.get(quest["id"])

        # Areas & categories for filters
        areas = [r[0] for r in conn.execute(
            "SELECT DISTINCT area FROM quests WHERE area IS NOT NULL AND area != '' ORDER BY area"
        ).fetchall()]
        categories = [r[0] for r in conn.execute(
            "SELECT DISTINCT difficulty FROM quests WHERE difficulty IS NOT NULL AND difficulty != '' ORDER BY difficulty"
        ).fetchall()]

    finally:
        conn.close()

    return {
        "quests": results,
        "total": total,
        "page": page,
        "per_page": per_page,
        "filters": {"areas": areas, "categories": categories},
    }
